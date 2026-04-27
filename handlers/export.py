"""Export handler — download loans data as CSV/JSON/Excel files."""
from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
from datetime import datetime, timezone, timedelta

from aiogram import Router, F
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    BufferedInputFile,
)

from bot.services.base.access import is_allowed
from bot.services.export.service import collect_export_entries

log = logging.getLogger(__name__)
router = Router(name="export")

SERVICES = {
    "kapusta": "🥬 Kapusta",
    "finkit": "🔵 FinKit",
    "zaimis": "🟪 ЗАЙМись",
    "mongo": "🦊 Mongo",
}

# Rate limiting: last export timestamps per key
_last_export: dict[str, datetime] = {}
RATE_LIMIT_ALL = timedelta(minutes=30)
RATE_LIMIT_SERVICE = timedelta(minutes=10)

# CSV columns for export
CSV_COLUMNS = [
    "id", "service", "request_type", "amount", "period_days",
    "interest_day", "interest_year", "penalty_interest", "credit_score",
    "created_at", "profit_gross", "profit_net", "amount_return",
    "platform_fee_open", "platform_fee_close", "full_name", "document_id",
    "display_name", "current_display_name", "display_names", "is_income_confirmed", "is_employed", "has_active_loan",
    "has_overdue", "note", "status", "loans_count",
    "opi_checked", "opi_has_debt", "opi_debt_amount", "opi_full_name", "opi_error",
    "kb_known", "kb_total_loans", "kb_settled", "kb_overdue", "kb_cancelled",
    "kb_has_claims", "kb_avg_rating", "kb_last_rating", "kb_total_invested",
    "loan_status_details_json", "enrichment_source", "source_account_tag",
]


@router.callback_query(F.data == "export_menu")
async def cb_export_menu(callback: CallbackQuery):
    if not await is_allowed(callback.message.chat.id):
        return

    lines = ["<b>📥 Скачать данные</b>\n"]
    lines.append("<i>Данные загружаются свежие с каждого сервиса.\n"
                 "FinKit: обогащение ПДФ + ОПИ + история заёмщиков.</i>")

    buttons = [
        [
            InlineKeyboardButton(text="🥬 Kapusta", callback_data="exp_fmt_kapusta"),
            InlineKeyboardButton(text="🔵 FinKit", callback_data="exp_fmt_finkit"),
        ],
        [
            InlineKeyboardButton(text="🟪 ЗАЙМись", callback_data="exp_fmt_zaimis"),
            InlineKeyboardButton(text="🦊 Mongo", callback_data="exp_fmt_mongo"),
        ],
        [InlineKeyboardButton(text="📁 Все — CSV", callback_data="exp_do_all_csv")],
        [InlineKeyboardButton(text="📋 Все — JSON", callback_data="exp_do_all_json")],
        [InlineKeyboardButton(text="📊 Все — Excel", callback_data="exp_do_all_xlsx")],
        [InlineKeyboardButton(text="↩ Главное меню", callback_data="main_menu")],
    ]
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await callback.message.edit_text("\n".join(lines), reply_markup=kb, parse_mode="HTML")


# ---- Per-service format selection ----

@router.callback_query(F.data.startswith("exp_fmt_"))
async def cb_export_format_select(callback: CallbackQuery):
    if not await is_allowed(callback.message.chat.id):
        return

    svc = callback.data.replace("exp_fmt_", "")
    label = SERVICES.get(svc, svc)

    buttons = [
        [InlineKeyboardButton(text="📁 CSV", callback_data=f"exp_do_{svc}_csv")],
        [InlineKeyboardButton(text="📋 JSON", callback_data=f"exp_do_{svc}_json")],
        [InlineKeyboardButton(text="📊 Excel", callback_data=f"exp_do_{svc}_xlsx")],
        [InlineKeyboardButton(text="↩ Назад", callback_data="export_menu")],
    ]
    kb = InlineKeyboardMarkup(inline_keyboard=buttons)
    await callback.message.edit_text(
        f"<b>📥 Формат выгрузки для {label}</b>",
        reply_markup=kb,
        parse_mode="HTML",
    )


def _entries_to_csv(entries: list[dict]) -> bytes:
    """Convert entries to CSV bytes (UTF-8 with BOM for Excel compatibility)."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for entry in entries:
        row = {k: entry.get(k, "") for k in CSV_COLUMNS}
        writer.writerow(row)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _entries_to_json(entries: list[dict]) -> bytes:
    """Convert entries to pretty JSON bytes."""
    return json.dumps(entries, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def _entries_to_xlsx(entries: list[dict]) -> bytes:
    """Convert entries to Excel bytes."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _entries_to_csv(entries)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    for col_idx, col_name in enumerate(CSV_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, entry in enumerate(entries, 2):
        for col_idx, col_name in enumerate(CSV_COLUMNS, 1):
            val = entry.get(col_name, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(CSV_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


FORMAT_HANDLERS = {
    "csv": (_entries_to_csv, "csv", "📁"),
    "json": (_entries_to_json, "json", "📋"),
    "xlsx": (_entries_to_xlsx, "xlsx", "📊"),
}


# ---- Unified export handler ----

@router.callback_query(F.data.startswith("exp_do_"))
async def cb_export_do(callback: CallbackQuery):
    if not await is_allowed(callback.message.chat.id):
        return

    # Parse: exp_do_{service}_{format}  or  exp_do_all_{format}
    parts = callback.data.replace("exp_do_", "").rsplit("_", 1)
    if len(parts) != 2:
        await callback.answer("❌ Неверный формат")
        return

    target, fmt = parts
    services = list(SERVICES.keys()) if target == "all" else [target]
    label = "все_сервисы" if target == "all" else target

    if fmt not in FORMAT_HANDLERS:
        await callback.answer("❌ Неизвестный формат")
        return

    # Rate limiting
    now = datetime.now(timezone.utc)
    rate_key = target  # "all" or service name
    limit = RATE_LIMIT_ALL if target == "all" else RATE_LIMIT_SERVICE
    last = _last_export.get(rate_key)
    if last and (now - last) < limit:
        remaining = limit - (now - last)
        mins = int(remaining.total_seconds() // 60) + 1
        await callback.answer(f"⏳ Подождите ≈{mins} мин", show_alert=True)
        return

    converter, ext, icon = FORMAT_HANDLERS[fmt]

    await callback.message.edit_text(f"⏳ Загрузка данных ({label}, {ext})...")

    try:
        entries = await collect_export_entries(services, callback.message.chat.id)
    except Exception as e:
        await callback.message.edit_text(
            f"❌ Ошибка загрузки: {e}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="↩ Назад", callback_data="export_menu")],
            ]),
        )
        return

    if not entries:
        await callback.message.edit_text(
            f"📭 Нет данных для {label}.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="↩ Назад", callback_data="export_menu")],
            ]),
        )
        return

    _last_export[rate_key] = now  # record successful export time

    file_bytes = converter(entries)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    filename = f"loans_{label}_{ts}.{ext}"

    # Count enriched stats
    kb_count = sum(1 for e in entries if e.get("kb_known"))
    opi_count = sum(1 for e in entries if e.get("opi_checked"))
    opi_err = sum(1 for e in entries if e.get("opi_error"))

    caption = f"{icon} {len(entries)} заявок ({label})"
    if kb_count:
        caption += f"\n👤 Известных заёмщиков: {kb_count}"
    if opi_count:
        caption += f"\n🔍 ОПИ проверено: {opi_count}"
    if opi_err:
        caption += f" (⚠️ ошибок: {opi_err})"

    await callback.message.answer_document(
        BufferedInputFile(file_bytes, filename=filename),
        caption=caption,
    )
    await callback.answer()
